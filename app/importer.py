"""
Workbook parsing + export for Revenue tracker.

parse_workbook_bytes() reads any .xlsm/.xlsx with the Revenue_2026 layout:
  - On-Site sheet  : master data, rows>=3, C=resource name (blank => subtotal/skip),
                     A=country B=client D=role E=billing rate, J..BJ = 53 weekly hours
  - Off-Shore sheet: same rows/names, E=expense rate (used for the expense column)

build_workbook() regenerates a file with the same layout incl. formulas and a
rebuilt Dashboard sheet, so exports can be re-imported or opened in Excel.
"""
from __future__ import annotations

import datetime as dt
import io
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

WEEK_COL_START = 10   # J
WEEK_COL_END = 62     # BJ  (53 weeks)
FIRST_DATA_ROW = 3

HEADERS = ["Country", "Client Name", "Resoruce Name", "Role", "Rate",
           "Utlilization", "Total Hrs", "Total Cost", "Revenue"]
OFF_HEADERS = ["Country", "Client Name", "Resoruce Name", "Role", "Offshore Rate",
               "Utlilization", "Total Hrs", "Total Cost", "Revenue"]

EU_COUNTRIES = {
    "EUROPE", "EU", "EUR", "FRANCE", "GERMANY", "SPAIN", "ITALY", "NETHERLANDS",
    "BELGIUM", "AUSTRIA", "PORTUGAL", "IRELAND", "FINLAND", "GREECE", "POLAND",
    "SWEDEN", "DENMARK", "NORWAY", "SWITZERLAND", "UK", "FR", "DE", "ES", "IT",
    "NL", "BE", "AT", "PT", "IE", "FI", "GR", "PL", "SE", "DK", "NO", "CH",
}


def default_layout(year: int) -> tuple[list[str], list[dict]]:
    """Generate J..BJ week labels matching the Revenue 2026 layout:
    first label = Jan-02 (year-start short week), then the first Monday of the
    year and every Monday after — 53 weeks total."""
    start = dt.date(year, 1, 2)
    weeks = [start.strftime("%b-%d")]
    cur = start + dt.timedelta(days=(7 - start.weekday()) % 7 or 7)  # first Monday after Jan-02
    while len(weeks) < 53:
        weeks.append(cur.strftime("%b-%d"))
        cur += dt.timedelta(days=7)
    months = []
    for w in weeks:
        m = w[:3].upper()
        if not months or months[-1]["name"] != m:
            months.append({"name": m, "start": len(weeks) - 1, "end": len(weeks) - 1})
        else:
            months[-1]["end"] = len(weeks) - 1
    return weeks, months


# Map of accepted header texts -> field (lowercased, trimmed)
KNOWN_COLS = {
    "country": "country",
    "client": "client", "client name": "client",
    "project": "project",
    "name": "name", "resource": "name", "resource name": "name", "resoruce name": "name",
    "role": "role", "title": "role",
    "rate": "rate", "onsite rate": "rate", "on-site rate": "rate", "billing rate": "rate",
    "offshore rate": "offshore_rate", "off-shore rate": "offshore_rate", "offshore": "offshore_rate",
}
SKIP_COLS = {
    "", "utilization", "utlilization", "total hrs", "total hours", "total cost",
    "total revenue", "revenue", "subtotal", "difference", "expense", "profit",
    "diff", "total", "currency", "cur", "action", "note", "remarks", "comments",
    "summary", "shift",
}


def _find_header_row(ws) -> tuple[int, dict, list[int]]:
    """Locate the header row (contains 'country' + a name-ish column), the
    column map, and the week column indexes. Returns (-1, {}, []) if absent."""
    for r in range(1, min(ws.max_row, 5) + 1):
        vals = [(c, str(ws.cell(r, c).value or "").strip().lower()) for c in range(1, ws.max_column + 1)]
        texts = [v[1] for v in vals]
        if "country" not in texts:
            continue
        if not any("name" in t or "role" in t or "title" in t or "resource" in t for t in texts):
            continue
        colmap: dict[str, int] = {}
        weeks: list[int] = []
        for c, t in vals:
            if t in KNOWN_COLS:
                colmap[KNOWN_COLS[t]] = c
            elif t in SKIP_COLS:
                continue
            elif t:
                weeks.append(c)
        if "name" not in colmap or not weeks:
            continue
        return r, colmap, weeks
    return -1, {}, []


def _month_bands(ws, header_row: int, week_cols: list[int], weeks: list[str]) -> list[dict]:
    """Month band labels live in the row above the header (row 1 of the classic
    layout). Fall back to deriving bands from the week label prefixes."""
    months: list[dict] = []
    if header_row > 1:
        for i, c in enumerate(week_cols):
            label = str(ws.cell(header_row - 1, c).value or "").strip().upper()
            if label and (not months or months[-1]["name"] != label):
                months.append({"name": label, "start": i, "end": i})
            elif months:
                months[-1]["end"] = i
    if not months:
        for i, w in enumerate(weeks):
            m = (w[:3] if w else "").upper()
            if m and (not months or months[-1]["name"] != m):
                months.append({"name": m, "start": i, "end": i})
            elif months and m:
                months[-1]["end"] = i
    return months


def _legacy_weeks(ws) -> tuple[list[str], list[int]]:
    weeks = []
    cols = list(range(WEEK_COL_START, WEEK_COL_END + 1))
    for c in cols:
        v = ws.cell(2, c).value
        if isinstance(v, dt.datetime):
            v = v.strftime("%b-%d")
        weeks.append(str(v or ""))
    return weeks, cols


def _as_float(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return 0.0
    return 0.0


def _norm(s) -> str:
    return (s or "").strip().upper()


def parse_workbook_bytes(data: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    if "On-Site" not in wb.sheetnames:
        raise ValueError("Workbook has no 'On-Site' sheet — expected the Revenue layout")

    ws = wb["On-Site"]
    header_row, colmap, week_cols = _find_header_row(ws)
    if header_row < 1:
        # Legacy fallback: fixed layout, headers on row 2, weeks J..BJ
        header_row = 2
        colmap = {"country": 1, "client": 2, "name": 3, "role": 4, "rate": 5}
        week_cols = list(range(WEEK_COL_START, WEEK_COL_END + 1))

    weeks = [str(ws.cell(header_row, c).value or "") for c in week_cols]
    if isinstance(ws.cell(header_row, week_cols[0]).value, dt.datetime):
        weeks = [v.strftime("%b-%d") if isinstance(v, dt.datetime) else v for v in weeks]
    if not weeks or not weeks[0]:
        raise ValueError("Could not read weekly columns from the header row")

    months = _month_bands(ws, header_row, week_cols, weeks)

    # Off-Shore rate source: header-mapped or legacy E column, keyed by row + name
    off_cols: dict = {}
    off_header_row = -1
    off_ws = None
    for sheet_name in ("Off-Shore", "Offshore"):
        if sheet_name in wb.sheetnames:
            off_ws = wb[sheet_name]
            break
    off_by_row: dict[int, float] = {}
    off_by_key: dict[tuple, float] = {}
    if off_ws is not None:
        off_header_row, off_cols, _off_weeks = _find_header_row(off_ws)
        off_rate_col = off_cols.get("offshore_rate") or (5 if off_header_row < 1 else None)
        if off_rate_col:
            data_start = (off_header_row + 1) if off_header_row > 0 else 3
            off_name_col = off_cols.get("name", 3)
            off_client_col = off_cols.get("client", 2)
            for r in range(data_start, off_ws.max_row + 1):
                nm = off_ws.cell(r, off_name_col).value
                if nm:
                    rate = off_ws.cell(r, off_rate_col).value
                    if isinstance(rate, (int, float)):
                        off_by_row[r] = round(float(rate), 4)
                        off_by_key[(_norm(off_ws.cell(r, off_client_col).value), _norm(nm))] = round(float(rate), 4)

    resources = []
    warnings = []
    name_col = colmap["name"]
    client_col = colmap.get("client", 2)
    project_col = colmap.get("project")
    role_col = colmap.get("role")
    rate_col = colmap.get("rate")
    country_col = colmap.get("country", 1)

    def _cell(r, c):
        if c is None:
            return None
        return ws.cell(r, c).value

    for r in range(header_row + 1, ws.max_row + 1):
        name = _cell(r, name_col)
        if not name or not str(name).strip():
            continue  # subtotal rows / blanks
        country = str(_cell(r, country_col) or "").strip()
        client = str(_cell(r, client_col) or "").strip()
        project = str(_cell(r, project_col) or "").strip() if project_col else ""
        role = str(_cell(r, role_col) or "").strip() if role_col else ""
        rate = _cell(r, rate_col)
        rate = float(rate) if isinstance(rate, (int, float)) else (float(rate) if isinstance(rate, str) and rate.strip() else None)
        if rate is not None:
            rate = round(rate, 4)

        hours = [_as_float(ws.cell(r, c).value) for c in week_cols]

        off_rate = off_by_row.get(r)
        if off_rate is None:
            off_rate = off_by_key.get((_norm(client), _norm(name)))
        if off_rate is None and name:
            warnings.append(f"row {r}: no Off-Shore rate for '{name}' — expense 0")

        resources.append({
            "country": country, "client": client, "project": project,
            "name": str(name).strip(), "role": role, "rate": rate,
            "offshore_rate": off_rate, "hours": hours,
        })
    wb.close()
    return {
        "resources": resources, "weeks": weeks, "months": months,
        "warnings": warnings, "has_project": bool(project_col),
    }


def parse_pricing_sheet(data: bytes) -> list[dict]:
    """Optional 'Pricing' sheet (Title / Rate / Offshore Rate / Currency) from an export."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return []
    if "Pricing" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Pricing"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        def _f(v):
            if isinstance(v, (int, float)):
                return round(float(v), 4)
            if isinstance(v, str) and v.strip():
                try:
                    return round(float(v.strip()), 4)
                except ValueError:
                    return None
            return None
        currency = "USD"
        if len(row) > 3 and row[3]:
            s = str(row[3]).strip().upper().replace("$", "USD").replace("£", "GBP")
            if s in ("USD", "GBP", "CAD"):
                currency = s
        out.append({"title": str(row[0]).strip(), "rate": _f(row[1]),
                    "offshore_rate": _f(row[2]) if len(row) > 2 else None,
                    "currency": currency})
    wb.close()
    return out


# ---------------- Export ----------------
def build_workbook(weeks: list[str], months: list[dict],
                   resources: list[dict], dashboard: dict,
                   pricing: list[dict] | None = None) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # --- On-Site / Off-Shore sheets ---
    _build_timesheet_sheet(wb, "On-Site", weeks, months, resources, billed=True)
    _build_timesheet_sheet(wb, "Off-Shore", weeks, months, resources, billed=False)
    # protect Off-Shore like the original (password from the VBA const)
    ws_off = wb["Off-Shore"]
    ws_off.protection.sheet = True
    ws_off.protection.password = "offshore2024"

    # --- Dashboard sheet ---
    ws_d = wb.active
    ws_d.title = "Dashboard"
    headers = ["Country", "Client", "Revenue", "Expense", "Difference", "Cur"]
    for c, h in enumerate(headers, start=1):
        cell = ws_d.cell(1, c, h)
        cell.font = Font(bold=True)
    bold = Font(bold=True)
    money = "#,##0.00"
    r = 2
    for g in dashboard["groups"]:
        ws_d.cell(r, 1, g["country"])
        ws_d.cell(r, 2, g["client"])
        ws_d.cell(r, 3, g["revenue"]).number_format = money
        ws_d.cell(r, 4, g["expense"]).number_format = money
        ws_d.cell(r, 5, g["difference"]).number_format = money
        ws_d.cell(r, 6, g["currency"])
        r += 1
    for t in dashboard["totals"]:
        ws_d.cell(r, 1, f"TOTAL {t['currency']}")
        ws_d.cell(r, 3, t["revenue"]).number_format = money
        ws_d.cell(r, 4, t["expense"]).number_format = money
        ws_d.cell(r, 5, t["difference"]).number_format = money
        ws_d.cell(r, 6, t["currency"])
        for c in range(1, 7):
            ws_d.cell(r, c).font = bold
        r += 1
    for col, width in zip("ABCDEF", (14, 22, 14, 14, 14, 8)):
        ws_d.column_dimensions[col].width = width

    # --- Pricing sheet (Title / Rate / Offshore Rate / Currency) ---
    ws_p = wb.create_sheet("Pricing")
    for c, h in enumerate(("Title", "Rate", "Offshore Rate", "Currency"), start=1):
        cell = ws_p.cell(1, c, h)
        cell.font = Font(bold=True)
    seen_titles: set[str] = set()
    if pricing:
        src = [{"role": p["title"], "rate": p["rate"], "offshore_rate": p["offshore_rate"],
                "currency": p.get("currency", "USD")} for p in pricing]
    else:
        src = resources
    rp = 2
    for res in src:
        title = (res["role"] or "").strip()
        if not title or title in seen_titles:
            continue
        seen_titles.add(title)
        ws_p.cell(rp, 1, title)
        ws_p.cell(rp, 2, res["rate"]).number_format = "#,##0.00"
        ws_p.cell(rp, 3, res["offshore_rate"]).number_format = "#,##0.00"
        ws_p.cell(rp, 4, res.get("currency", "USD"))
        rp += 1
    for col, width in zip("ABCD", (34, 12, 14, 10)):
        ws_p.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf


def _build_timesheet_sheet(wb, title, weeks, months, resources, billed: bool) -> None:
    ws = wb.create_sheet(title)
    headers = (
        ["Country", "Client", "Project", "Resource Name", "Title", "Rate",
         "Total Hours", "Total Revenue"]
        if billed else
        ["Country", "Client", "Project", "Resource Name", "Title", "Offshore Rate",
         "Total Hours", "Total Revenue"]
    )
    WEEK0 = 9  # column I — weeks start here

    # row 1: month band labels (above the header, classic layout)
    for m in months:
        ws.cell(1, WEEK0 + m["start"], m["name"])

    # row 2: headers
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    for c, w in enumerate(weeks, start=WEEK0):
        ws.cell(2, c, w)

    end_col = WEEK0 + len(weeks) - 1
    end_letter = get_column_letter(end_col)
    r = 3
    for res in resources:
        ws.cell(r, 1, res["country"])
        ws.cell(r, 2, res["client"])
        ws.cell(r, 3, res["project"] or "")
        ws.cell(r, 4, res["name"])
        ws.cell(r, 5, res["role"])
        ws.cell(r, 6, res["rate"] if billed else res["offshore_rate"])
        ws.cell(r, 7, f"=SUM(I{r}:{end_letter}{r})")   # Total Hours
        ws.cell(r, 8, f"=F{r}*G{r}")                    # Total Revenue
        for i, h in enumerate(res["hours"]):
            if h:
                ws.cell(r, WEEK0 + i, h)
        r += 1
    for col, width in zip("ABCDEFGH", (12, 18, 16, 18, 22, 12, 12, 14)):
        ws.column_dimensions[col].width = width